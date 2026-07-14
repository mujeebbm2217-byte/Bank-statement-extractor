"""
Bank Statement Keyword Amount Extractor - Streamlit Version (Upgraded)
-----------------------------------------------------------------------------
Run locally with:   streamlit run app.py
Deploy free on:      https://share.streamlit.io  (Streamlit Community Cloud)
                      -- push this file + requirements.txt + packages.txt to
                      a GitHub repo, then "New app" on share.streamlit.io.

-----------------------------------------------------------------------------
WHAT'S NEW IN THIS VERSION (vs your original app.py)
-----------------------------------------------------------------------------
All your original features still work exactly as before:
    - Login gate
    - PDF -> Excel and Excel/CSV -> PDF converters
    - Keyword amount search with st.data_editor manual correction
    - Header row detection + column auto-mapping

On top of that, this version adds a much stronger OCR pipeline for poor
quality / scanned bank statements:

    1. Image preprocessing (OpenCV) before OCR:
       CLAHE contrast boost -> denoise -> deskew -> adaptive threshold ->
       morphological cleanup. This alone fixes a large % of OCR misreads.

    2. Higher default OCR resolution (450 DPI, with a 600 DPI "poor scan"
       option you can pick per-upload) instead of the old fixed 300 DPI.

    3. A 3-engine OCR fallback chain: PaddleOCR -> Tesseract -> EasyOCR.
       If the best engine (PaddleOCR) isn't installed/available, it quietly
       falls back to the next one -- your app never breaks, it just uses
       whichever engine it can.

    4. Per-row OCR confidence scores, with low-confidence rows flagged in
       the preview table so you know exactly which rows to double check.

    5. A bank/UPI keyword correction dictionary (BANK_WORDS) + fuzzy
       matching (rapidfuzz) to auto-fix common OCR misreads of payment
       app names, e.g. "GOOGIEPAY" -> "GOOGLEPAY".

    6. OCR digit-confusion correction (O->0, I->1, l->1, S->5, B->8) applied
       automatically to amount-looking OCR text before it's parsed as a
       number.

    7. A balance validation engine: Previous Balance + Credit - Debit should
       equal Current Balance. Any row that doesn't match is flagged as a
       possible OCR error, right in the app.

    8. Table extraction now tries, in order: pdfplumber -> Camelot -> OCR.
       Camelot is only used as a fallback when pdfplumber can't find a
       clean table on a page (it needs Ghostscript installed -- see
       packages.txt).

IMPORTANT DEPLOYMENT NOTE (read before pushing to Streamlit Cloud):
    PaddleOCR + PaddlePaddle + EasyOCR together are HEAVY (600 MB - 1 GB+ of
    model/library downloads). Streamlit Community Cloud's free tier has a
    1 GB RAM limit and can struggle to install/run all three. That's exactly
    why every engine in this file is wrapped in try/except and the app keeps
    working with whichever engines are actually available (worst case, it
    falls back to Tesseract only, same as your original app). If you hit
    memory errors on Community Cloud, the safest fix is to remove
    "paddleocr"/"paddlepaddle" from requirements.txt and keep just
    Tesseract + EasyOCR (or Tesseract only).

-----------------------------------------------------------------------------
IMPORTANT: LOGIN / PASSWORD NOTE (unchanged from your original file)
-----------------------------------------------------------------------------
Streamlit Community Cloud apps are PUBLIC by default (anyone with the link
can open them) unless you turn on Streamlit's built-in viewer auth in the
app settings. The username/password gate below is a lightweight extra layer.
We only store a SHA-256 fingerprint, never the real password.

To set your own password:
    1. Run this once in any terminal (no need to run the app):
           python -c "import hashlib; print(hashlib.sha256(b'yourNewPassword').hexdigest())"
    2. Copy the printed hash into APP_PASSWORD_HASH below.
"""

import hashlib
import io
import os
import re
import tempfile

import numpy as np
import pandas as pd
import pdfplumber
import openpyxl
from openpyxl.utils import get_column_letter
import streamlit as st
from PIL import Image

# ----------------------------- Optional libraries -----------------------------
# Every optional dependency is wrapped in try/except so the app NEVER crashes
# on import, even if a heavy library (PaddleOCR, EasyOCR, Camelot...) failed
# to install on the deployment server. Each _AVAILABLE flag is checked before
# that feature is used, and the UI tells the user what's on/off.

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
    from reportlab.lib.units import mm
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import pytesseract
    from pytesseract import Output as _TesseractOutput
    PYTESSERACT_AVAILABLE = True
except ImportError:
    PYTESSERACT_AVAILABLE = False
# ^ OCR needs BOTH the "pytesseract" Python package (in requirements.txt)
#   AND the Tesseract engine itself installed on the machine/server:
#     - Streamlit Community Cloud: add a file named "packages.txt"
#       (next to app.py) containing:  tesseract-ocr

try:
    import cv2
    OPENCV_AVAILABLE = True
except ImportError:
    OPENCV_AVAILABLE = False
# ^ Used for CLAHE / denoise / deskew / adaptive-threshold preprocessing.
#   If missing, OCR still runs on the raw page image, just less accurately.

try:
    from paddleocr import PaddleOCR
    PADDLEOCR_AVAILABLE = True
except ImportError:
    PADDLEOCR_AVAILABLE = False

try:
    import easyocr
    EASYOCR_AVAILABLE = True
except ImportError:
    EASYOCR_AVAILABLE = False

try:
    from rapidfuzz import fuzz, process as rf_process
    RAPIDFUZZ_AVAILABLE = True
except ImportError:
    RAPIDFUZZ_AVAILABLE = False

try:
    import camelot
    CAMELOT_AVAILABLE = True
except ImportError:
    CAMELOT_AVAILABLE = False
# ^ Camelot needs the Ghostscript system binary too -- add "ghostscript" to
#   packages.txt on Streamlit Cloud, or table extraction just skips this
#   step and falls back to pdfplumber/OCR as before.

ANY_OCR_ENGINE_AVAILABLE = PYTESSERACT_AVAILABLE or PADDLEOCR_AVAILABLE or EASYOCR_AVAILABLE

NONE_OPTION = "-- None --"
DEVELOPER_NAME = "Mohammad Mujeeb"

# ----------------------------- Theme (navy/gold) -----------------------------
NAVY_BG = "#0f172a"
GOLD = "#eab308"
TEAL = "#14b8a6"

APP_USERNAME = "admin"
APP_PASSWORD_HASH = "707cfae1e410d17ac820cd4b290040561bb963b00fa1c13807d63723539fdba0"
# ^ hash for "changeme123" -- change this, see note above.


def _hash_password(plain_text_password: str) -> str:
    return hashlib.sha256(plain_text_password.encode("utf-8")).hexdigest()


# ============================================================================
# BANK / UPI KEYWORD CORRECTION  (new)
# ============================================================================
# Common OCR misreads of payment-app / bank names seen in real statements.
# Add more entries here any time you spot a new recurring misread -- it's
# just a plain dictionary, no code changes needed elsewhere.
BANK_WORDS = {
    "GOOGIEPAY": "GOOGLEPAY",
    "GOOGLEPAYI": "GOOGLEPAY",
    "PAYTMFHDFC": "PAYTM-HDFC",
    "PHONPE": "PHONEPE",
    "UPL": "UPI",
    "UPIAAR": "UPIAR",
}

_BANK_WORD_FUZZY_THRESHOLD = 85  # 0-100, rapidfuzz similarity score cutoff


def correct_bank_keywords(text, threshold=_BANK_WORD_FUZZY_THRESHOLD):
    """Fix known OCR misreads of bank/UPI keywords in a Reference string.
    Step 1: exact whole-word replacement using BANK_WORDS.
    Step 2 (if rapidfuzz installed): fuzzy-match any leftover word against
    the BANK_WORDS keys and correct it if it's a close-enough match, so
    variants you haven't explicitly listed still get caught.
    """
    if text is None or str(text).strip() == '':
        return text
    corrected = str(text)

    # Exact match on whole words (case-insensitive)
    for wrong, right in BANK_WORDS.items():
        corrected = re.sub(r'\b' + re.escape(wrong) + r'\b', right, corrected, flags=re.IGNORECASE)

    if RAPIDFUZZ_AVAILABLE:
        tokens = corrected.split()
        known_correct_values = set(BANK_WORDS.values())
        fixed_tokens = []
        for tok in tokens:
            clean_tok = re.sub(r'[^A-Za-z]', '', tok)
            if len(clean_tok) < 4 or clean_tok.upper() in known_correct_values:
                fixed_tokens.append(tok)
                continue
            match = rf_process.extractOne(clean_tok.upper(), BANK_WORDS.keys(), scorer=fuzz.ratio)
            if match and match[1] >= threshold:
                fixed_tokens.append(BANK_WORDS[match[0]])
            else:
                fixed_tokens.append(tok)
        corrected = ' '.join(fixed_tokens)

    return corrected


# ============================================================================
# OCR DIGIT-CONFUSION CORRECTION  (new)
# ============================================================================
# Letters that Tesseract/PaddleOCR/EasyOCR frequently confuse with digits on
# low-quality scans. Applied ONLY to strings that already look numeric (i.e.
# amount / balance columns), never to free-text narration -- otherwise it
# would corrupt real words.
OCR_DIGIT_FIX = {'O': '0', 'o': '0', 'I': '1', 'l': '1', 'S': '5', 'B': '8'}


def fix_amount_ocr_chars(s):
    """Replace common letter/digit OCR confusions inside a numeric-looking
    string, e.g. 'S000' -> '5000', 'l2OO.OO' -> '1200.00'."""
    if s is None:
        return s
    s = str(s)
    stripped = s.strip()
    # Only touch strings made up of digits/commas/dot/minus plus the
    # confusable letters -- if there's any other letter, it's real text
    # (like a narration), so leave it alone.
    if stripped and re.fullmatch(r"[0-9OoIlSB,.\-\s]+", stripped):
        return ''.join(OCR_DIGIT_FIX.get(ch, ch) for ch in s)
    return s


# ----------------------------- Core parsing logic -----------------------------
# (Table-strategy scoring, wrapped-narration merge heuristic, column
# auto-mapping, amount cleaning -- unchanged in behaviour from your original
# file, with the OCR digit-fix hook added to clean_amount.)

PDF_TABLE_STRATEGIES = [
    {"vertical_strategy": "lines", "horizontal_strategy": "lines"},
    {"vertical_strategy": "lines", "horizontal_strategy": "text",
     "text_x_tolerance": 1, "text_y_tolerance": 3,
     "snap_tolerance": 3, "join_tolerance": 3},
    {"vertical_strategy": "text", "horizontal_strategy": "text",
     "text_x_tolerance": 1, "text_y_tolerance": 3},
]

_DATE_CELL_RE = re.compile(r'^\d{1,2}[/-]\d{1,2}[/-]\d{2,4}$')


def map_columns(columns):
    """Map actual column names to standard names: reference, date, debit,
    credit, balance. Only a STARTING GUESS -- user can override in the UI."""
    mapping = {}
    col_lower = {c: str(c).lower().strip() for c in columns}

    date_kw = ['date', 'txn date', 'value date', 'transaction date']
    ref_kw = ['narration', 'description', 'particulars', 'reference', 'remarks',
              'transaction details', 'details', 'transaction remarks', 'recipient', 'payee']
    debit_kw = ['debit', 'withdrawal', 'dr amount', 'withdrawal amt', 'payment', 'payment amt']
    credit_kw = ['credit', 'deposit', 'cr amount', 'deposit amt', 'receipt', 'receipt amt']
    balance_kw = ['balance', 'closing balance', 'running balance']

    for col, low in col_lower.items():
        if 'reference' not in mapping and any(k in low for k in ref_kw):
            mapping['reference'] = col
        elif 'date' not in mapping and any(k in low for k in date_kw):
            mapping['date'] = col
        elif 'debit' not in mapping and any(k in low for k in debit_kw):
            mapping['debit'] = col
        elif 'credit' not in mapping and any(k in low for k in credit_kw):
            mapping['credit'] = col
        elif 'balance' not in mapping and any(k in low for k in balance_kw):
            mapping['balance'] = col
    return mapping


def clean_amount(val, ocr_mode=False):
    """Convert amount string/number to float, handling commas, currency
    symbols, blanks. When ocr_mode=True, first fixes common OCR digit
    confusions (O->0, I/l->1, S->5, B->8) before parsing."""
    if pd.isna(val):
        return np.nan
    s = str(val).strip()
    if s == '' or s.lower() in ['nan', '-', 'none']:
        return np.nan
    if ocr_mode:
        s = fix_amount_ocr_chars(s)
    s = re.sub(r'[^\d.\-]', '', s)
    if s in ['', '-', '.']:
        return np.nan
    try:
        return float(s)
    except ValueError:
        return np.nan


def _is_blank_row(row):
    return all((c is None or str(c).strip() == '') for c in row)


def _score_strategy(tables):
    """Count rows containing a date-like cell -- a strong bank-statement
    -specific signal of a real transaction row."""
    score = 0
    for table in tables:
        for row in table:
            if not row:
                continue
            if any(c and _DATE_CELL_RE.match(str(c).strip()) for c in row):
                score += 1
    return score


def _extract_pdf_tables_best_strategy(pdf):
    """Try each strategy on a couple of sample pages and score how many real
    transaction rows each one recovers."""
    sample_pages = pdf.pages[:min(2, len(pdf.pages))]
    scores = []
    for settings in PDF_TABLE_STRATEGIES:
        total = 0
        for page in sample_pages:
            try:
                total += _score_strategy(page.extract_tables(settings))
            except Exception:
                pass
        scores.append(total)

    best_score = max(scores) if scores else 0
    threshold = best_score * 0.7
    for settings, score in zip(PDF_TABLE_STRATEGIES, scores):
        if score >= threshold:
            return settings
    return PDF_TABLE_STRATEGIES[0]


# ============================================================================
# IMAGE PREPROCESSING FOR OCR  (new)
# ============================================================================

def _deskew_gray(gray):
    """Estimate and correct page skew using the minimum-area bounding
    rectangle of foreground pixels. Returns the gray image unchanged if a
    skew angle can't be reliably estimated (avoids making things worse)."""
    try:
        inv = cv2.bitwise_not(gray)
        thresh = cv2.threshold(inv, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
        coords = np.column_stack(np.where(thresh > 0))
        if coords.shape[0] < 20:
            return gray
        angle = cv2.minAreaRect(coords)[-1]
        if angle < -45:
            angle = -(90 + angle)
        else:
            angle = -angle
        # Ignore negligible skew or clearly-wrong detections (full pages of
        # text rarely need more than a few degrees of correction).
        if abs(angle) < 0.1 or abs(angle) > 15:
            return gray
        (h, w) = gray.shape[:2]
        matrix = cv2.getRotationMatrix2D((w // 2, h // 2), angle, 1.0)
        rotated = cv2.warpAffine(
            gray, matrix, (w, h), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE
        )
        return rotated
    except Exception:
        return gray


def preprocess_for_ocr(pil_image):
    """Full preprocessing pipeline for poor-quality scanned bank statements:
        grayscale -> CLAHE contrast -> denoise -> deskew -> adaptive
        threshold -> morphological cleanup.
    Returns a PIL Image ready to hand to any OCR engine. If OpenCV isn't
    installed, returns the original image untouched (OCR will just be a bit
    less accurate on bad scans -- it still works)."""
    if not OPENCV_AVAILABLE:
        return pil_image
    try:
        img = cv2.cvtColor(np.array(pil_image.convert('RGB')), cv2.COLOR_RGB2BGR)
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

        # 1. CLAHE contrast enhancement -- boosts faint print on poor scans
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
        gray = clahe.apply(gray)

        # 2. Noise removal
        gray = cv2.fastNlMeansDenoising(gray, h=10)

        # 3. Deskewing
        gray = _deskew_gray(gray)

        # 4. Adaptive thresholding -- handles uneven lighting/scan shadows
        thresh = cv2.adaptiveThreshold(
            gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 31, 15
        )

        # 5. Morphological cleanup -- removes tiny speck noise
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 1))
        cleaned = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, kernel)

        return Image.fromarray(cleaned).convert('RGB')
    except Exception:
        # Preprocessing is a "nice to have" -- never let it break OCR.
        return pil_image


# ============================================================================
# MULTI-ENGINE OCR PIPELINE  (new)
# ============================================================================
# All three engines are normalised into the same "word list" format:
#   {'text': str, 'left': int, 'top': int, 'width': int, 'height': int, 'conf': float 0-100}
# so a single clustering function can turn any engine's output into table
# rows. This is the same idea your original ocr_image_to_rows() used for
# Tesseract, generalised to also work with PaddleOCR / EasyOCR.

def _tesseract_words(pil_image, min_confidence=40):
    if not PYTESSERACT_AVAILABLE:
        raise RuntimeError("pytesseract is not installed.")
    data = pytesseract.image_to_data(pil_image, output_type=_TesseractOutput.DICT)
    words = []
    for i in range(len(data['text'])):
        text = data['text'][i].strip()
        if not text:
            continue
        try:
            conf = float(data['conf'][i])
        except (ValueError, TypeError):
            conf = -1
        if conf < 0:
            continue
        words.append({
            'text': text, 'left': data['left'][i], 'top': data['top'][i],
            'width': data['width'][i], 'height': data['height'][i], 'conf': conf,
        })
    return [w for w in words if w['conf'] >= min_confidence]


@st.cache_resource(show_spinner=False)
def _get_paddle_engine():
    return PaddleOCR(use_angle_cls=True, lang='en', show_log=False)


def _paddleocr_words(pil_image, min_confidence=40):
    if not PADDLEOCR_AVAILABLE:
        raise RuntimeError("paddleocr is not installed.")
    engine = _get_paddle_engine()
    img_arr = np.array(pil_image.convert('RGB'))
    result = engine.ocr(img_arr, cls=True)
    words = []
    if result and result[0]:
        for box, (text, conf) in result[0]:
            text = (text or '').strip()
            if not text:
                continue
            xs = [p[0] for p in box]
            ys = [p[1] for p in box]
            left, top = min(xs), min(ys)
            width, height = max(xs) - left, max(ys) - top
            conf_pct = conf * 100
            if conf_pct >= min_confidence:
                words.append({'text': text, 'left': left, 'top': top,
                               'width': width, 'height': height, 'conf': conf_pct})
    return words


@st.cache_resource(show_spinner=False)
def _get_easyocr_reader():
    return easyocr.Reader(['en'], gpu=False)


def _easyocr_words(pil_image, min_confidence=40):
    if not EASYOCR_AVAILABLE:
        raise RuntimeError("easyocr is not installed.")
    reader = _get_easyocr_reader()
    img_arr = np.array(pil_image.convert('RGB'))
    results = reader.readtext(img_arr)
    words = []
    for box, text, conf in results:
        text = (text or '').strip()
        if not text:
            continue
        xs = [p[0] for p in box]
        ys = [p[1] for p in box]
        left, top = min(xs), min(ys)
        width, height = max(xs) - left, max(ys) - top
        conf_pct = conf * 100
        if conf_pct >= min_confidence:
            words.append({'text': text, 'left': left, 'top': top,
                           'width': width, 'height': height, 'conf': conf_pct})
    return words


def _cluster_words_into_rows(words, column_gap_px=25, line_tol_ratio=0.6):
    """Group a flat list of positioned words into table rows/cells, using
    vertical position for rows and horizontal gaps for column boundaries.
    Returns (rows, row_confidences) where row_confidences[i] is the average
    OCR confidence (0-100) of every word in that row."""
    if not words:
        return [], []

    words = sorted(words, key=lambda w: (w['top'], w['left']))
    lines = []
    current_line = [words[0]]
    current_top = words[0]['top']
    current_height = words[0]['height'] or 20

    for w in words[1:]:
        tol = max(current_height, w['height'] or 20) * line_tol_ratio
        if abs(w['top'] - current_top) <= tol:
            current_line.append(w)
        else:
            lines.append(current_line)
            current_line = [w]
            current_top = w['top']
            current_height = w['height'] or 20
    lines.append(current_line)

    rows, row_confidences = [], []
    for line_words in lines:
        line_words = sorted(line_words, key=lambda w: w['left'])
        cells, confs = [], []
        current = [line_words[0]]
        for prev_w, cur_w in zip(line_words, line_words[1:]):
            gap = cur_w['left'] - (prev_w['left'] + prev_w['width'])
            if gap > column_gap_px:
                cells.append(' '.join(w['text'] for w in current))
                confs.extend(w['conf'] for w in current)
                current = [cur_w]
            else:
                current.append(cur_w)
        cells.append(' '.join(w['text'] for w in current))
        confs.extend(w['conf'] for w in current)
        rows.append(cells)
        row_confidences.append(sum(confs) / len(confs) if confs else 0.0)

    return rows, row_confidences


def run_ocr_pipeline(pil_image, min_confidence=40, column_gap_px=25, preprocess=True):
    """The full OCR pipeline requested:
        preprocess -> PaddleOCR -> (if failed) Tesseract -> (if failed) EasyOCR
    Returns (rows, row_confidences, engine_used). Raises RuntimeError only if
    every available engine failed AND none are installed / none produced
    output -- callers should catch this and fall back gracefully.
    """
    image = preprocess_for_ocr(pil_image) if preprocess else pil_image

    engine_chain = [
        ('PaddleOCR', PADDLEOCR_AVAILABLE, _paddleocr_words),
        ('Tesseract', PYTESSERACT_AVAILABLE, _tesseract_words),
        ('EasyOCR', EASYOCR_AVAILABLE, _easyocr_words),
    ]

    last_error = None
    for name, available, fn in engine_chain:
        if not available:
            continue
        try:
            words = fn(image, min_confidence)
            if not words:
                continue
            rows, row_confidences = _cluster_words_into_rows(words, column_gap_px)
            if rows:
                return rows, row_confidences, name
        except Exception as e:
            last_error = e
            continue

    if last_error is not None:
        raise RuntimeError(f"All OCR engines failed. Last error: {last_error}")
    raise RuntimeError(
        "No OCR engine is installed. Add at least one of pytesseract / "
        "paddleocr / easyocr to requirements.txt."
    )


# ============================================================================
# TABLE EXTRACTION: pdfplumber -> Camelot -> OCR  (new ordering)
# ============================================================================

def _camelot_extract_page(pdf_path, page_num, min_rows=2):
    """Try Camelot on a single page (lattice first, then stream). Returns a
    list of rows, or None if Camelot isn't available / didn't find a usable
    table -- caller then falls back to OCR."""
    if not CAMELOT_AVAILABLE:
        return None
    for flavor in ('lattice', 'stream'):
        try:
            tables = camelot.read_pdf(pdf_path, pages=str(page_num), flavor=flavor)
            if tables and tables.n > 0:
                df = tables[0].df
                rows = df.values.tolist()
                real_rows = [r for r in rows if not _is_blank_row(r)]
                if len(real_rows) >= min_rows:
                    return rows
        except Exception:
            continue
    return None


def page_needs_ocr(tables, min_rows=2):
    """A page 'needs OCR' if pdfplumber's normal table extraction found
    little or nothing -- the classic sign of a scanned/image-only page."""
    if not tables:
        return True
    main_table = max(tables, key=len)
    real_rows = [r for r in main_table if r and not _is_blank_row(r)]
    return len(real_rows) < min_rows


def extract_raw_rows(uploaded_file, ext, use_ocr=True, force_ocr=False,
                      ocr_resolution=450, apply_bank_correction=False):
    """
    Extract RAW rows (no header assumption yet) plus optional PDF page
    images for preview. `uploaded_file` is a Streamlit UploadedFile
    (file-like / BytesIO) -- pdfplumber and pandas both accept it directly.

    Extraction order per PDF page: pdfplumber -> Camelot -> OCR.
    Every row also gets a parallel confidence score (100 for digitally
    extracted rows, the OCR engine's own confidence for OCR'd rows) which is
    stashed in st.session_state["row_confidences"] for the UI to highlight.

    Returns (rows, page_images).
    """
    page_images = []
    ocr_pages_used = []       # page numbers that fell back to OCR
    ocr_engine_per_page = {}  # page_num -> engine name, for the UI message
    confidences = []          # parallel list to `rows`

    if ext == 'pdf':
        rows = []
        uploaded_file.seek(0)
        file_bytes = uploaded_file.read()
        uploaded_file.seek(0)

        # Camelot needs a real file path, not a stream -- write one temp copy.
        tmp_pdf_path = None
        if CAMELOT_AVAILABLE:
            try:
                tmp_fd, tmp_pdf_path = tempfile.mkstemp(suffix='.pdf')
                with os.fdopen(tmp_fd, 'wb') as f:
                    f.write(file_bytes)
            except Exception:
                tmp_pdf_path = None

        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            settings = _extract_pdf_tables_best_strategy(pdf)

            for page_num, page in enumerate(pdf.pages, start=1):
                try:
                    tables = page.extract_tables(settings)
                except Exception:
                    tables = []

                needs_fallback = force_ocr or page_needs_ocr(tables)

                if not needs_fallback and tables:
                    # 1) pdfplumber succeeded -- digital text, full confidence.
                    main_table = max(tables, key=len)
                    for r in main_table:
                        if r is None or _is_blank_row(r):
                            continue
                        rows.append(list(r))
                        confidences.append(100.0)
                    try:
                        img = page.to_image(resolution=100).original
                        page_images.append(img)
                    except Exception:
                        pass
                    continue

                # 2) pdfplumber gave nothing usable -- try Camelot next.
                camelot_rows = None
                if tmp_pdf_path is not None and not force_ocr:
                    camelot_rows = _camelot_extract_page(tmp_pdf_path, page_num)

                if camelot_rows is not None:
                    for r in camelot_rows:
                        if _is_blank_row(r):
                            continue
                        rows.append(list(r))
                        confidences.append(100.0)
                    try:
                        img = page.to_image(resolution=100).original
                        page_images.append(img)
                    except Exception:
                        pass
                    continue

                # 3) Last resort -- OCR (only if the user enabled it).
                if use_ocr and ANY_OCR_ENGINE_AVAILABLE:
                    try:
                        ocr_img = page.to_image(resolution=ocr_resolution).original
                        ocr_rows, ocr_confs, engine_used = run_ocr_pipeline(ocr_img)
                        if ocr_rows:
                            rows.extend(ocr_rows)
                            confidences.extend(ocr_confs)
                            ocr_pages_used.append(page_num)
                            ocr_engine_per_page[page_num] = engine_used
                        elif tables:
                            main_table = max(tables, key=len)
                            for r in main_table:
                                if r is None or _is_blank_row(r):
                                    continue
                                rows.append(list(r))
                                confidences.append(50.0)
                    except Exception:
                        # Every OCR engine missing/failed -- fall back
                        # silently to whatever pdfplumber gave us rather
                        # than losing the whole page.
                        if tables:
                            main_table = max(tables, key=len)
                            for r in main_table:
                                if r is None or _is_blank_row(r):
                                    continue
                                rows.append(list(r))
                                confidences.append(50.0)
                elif tables:
                    main_table = max(tables, key=len)
                    for r in main_table:
                        if r is None or _is_blank_row(r):
                            continue
                        rows.append(list(r))
                        confidences.append(50.0)

                try:
                    img = page.to_image(resolution=100).original
                    page_images.append(img)
                except Exception:
                    pass

        if tmp_pdf_path is not None:
            try:
                os.remove(tmp_pdf_path)
            except Exception:
                pass

        st.session_state["_ocr_pages_used"] = ocr_pages_used
        st.session_state["_ocr_engine_per_page"] = ocr_engine_per_page
        st.session_state["row_confidences"] = confidences
        st.session_state["_ocr_was_used"] = len(ocr_pages_used) > 0

        if apply_bank_correction:
            rows = [[correct_bank_keywords(c) if isinstance(c, str) else c for c in r] for r in rows]

        return rows, page_images

    uploaded_file.seek(0)
    if ext == 'csv':
        df_raw = pd.read_csv(uploaded_file, header=None, dtype=str)
    elif ext in ('xlsx', 'xls'):
        df_raw = pd.read_excel(uploaded_file, header=None, dtype=str)
    else:
        raise ValueError(f"Unsupported file type: {ext}")

    df_raw = df_raw.fillna('')
    rows = df_raw.values.tolist()
    st.session_state["row_confidences"] = [100.0] * len(rows)
    st.session_state["_ocr_was_used"] = False
    st.session_state["_ocr_pages_used"] = []
    return rows, page_images


def pad_row(row, width, merge_into=None):
    """Force a row to exactly `width` cells. Overflow cells (extra cells from
    wrapped narration text) are merged into `merge_into` instead of dropped."""
    row = ['' if c is None else c for c in row]

    if len(row) < width:
        row = row + [''] * (width - len(row))
    elif len(row) > width:
        target = merge_into if merge_into is not None else width - 1
        target = max(0, min(target, width - 1))

        overflow = row[width:]
        kept = row[:width]

        pieces = [str(kept[target])] + [str(c) for c in overflow if str(c).strip() != '']
        kept[target] = ' '.join(p for p in pieces if p.strip() != '')

        row = kept
    return row


def merge_wrapped_continuation_rows(rows, date_idx=None, debit_idx=None,
                                     credit_idx=None, narration_idx=None,
                                     confidences=None):
    """Fold "orphan" wrapped-narration rows into the transaction row above
    them. If `confidences` is given (parallel list to `rows`), it is merged
    the same way (min of the two rows) so confidence tracking survives the
    merge."""
    merged = []
    merged_confs = [] if confidences is not None else None

    for idx, row in enumerate(rows):
        row = list(row)
        conf = confidences[idx] if confidences is not None and idx < len(confidences) else 100.0

        def cell(i):
            return row[i] if 0 <= i < len(row) and row[i] is not None else ''

        has_any_text = any(str(c).strip() != '' for c in row)

        is_continuation = False
        if merged and has_any_text:
            checks = []
            if date_idx is not None:
                checks.append(str(cell(date_idx)).strip() == '')
            if debit_idx is not None:
                checks.append(str(cell(debit_idx)).strip() == '')
            if credit_idx is not None:
                checks.append(str(cell(credit_idx)).strip() == '')
            if checks:
                is_continuation = all(checks)
            else:
                is_continuation = str(cell(0)).strip() == ''

        if is_continuation:
            extra_text = ' '.join(str(c).strip() for c in row if str(c).strip() != '')
            prev = merged[-1]
            target = narration_idx if narration_idx is not None else (len(prev) - 1)
            target = max(0, min(target, len(prev) - 1)) if prev else 0
            if prev:
                existing = str(prev[target]) if prev[target] is not None else ''
                prev[target] = (existing + ' ' + extra_text).strip() if existing else extra_text
                if merged_confs is not None:
                    merged_confs[-1] = min(merged_confs[-1], conf)
            else:
                merged.append(row)
                if merged_confs is not None:
                    merged_confs.append(conf)
        else:
            merged.append(row)
            if merged_confs is not None:
                merged_confs.append(conf)

    if confidences is not None:
        return merged, merged_confs
    return merged


def build_clean_df(df_raw, col_map, ocr_mode=False):
    """col_map keys used: 'reference' (required), 'debit' (optional),
    'credit' (optional), 'balance' (optional, new). `ocr_mode` triggers the
    O/I/l/S/B digit-confusion fix before parsing amounts."""
    df = pd.DataFrame()
    df['Reference'] = df_raw[col_map['reference']].astype(str)

    if 'debit' in col_map:
        df['Debit'] = df_raw[col_map['debit']].apply(lambda v: clean_amount(v, ocr_mode=ocr_mode))
    if 'credit' in col_map:
        df['Credit'] = df_raw[col_map['credit']].apply(lambda v: clean_amount(v, ocr_mode=ocr_mode))
    if 'balance' in col_map:
        df['Balance'] = df_raw[col_map['balance']].apply(lambda v: clean_amount(v, ocr_mode=ocr_mode))

    df = df[df['Reference'].str.strip().str.lower().apply(lambda x: x not in ['', 'nan', 'none'])]
    df = df.reset_index(drop=True)
    return df


def guess_header_row(rows, max_scan=40):
    """Best-effort guess at which row is the real header row, so the UI can
    default the row-picker to something sensible instead of always 0."""
    keywords = ['date', 'narration', 'description', 'particulars', 'reference',
                'debit', 'credit', 'withdrawal', 'deposit', 'balance']
    for idx, row in enumerate(rows[:max_scan]):
        text = ' '.join(str(c).lower() for c in row if c is not None)
        if sum(k in text for k in keywords) >= 2:
            return idx
    return 0


# ============================================================================
# BALANCE VALIDATION ENGINE  (new)
# ============================================================================

def validate_balances(df, tolerance=1.0):
    """Checks: Previous Balance + Credit - Debit == Current Balance for every
    row. Returns (df_with_check_column, mismatch_count). Rows where Balance
    is missing, or where there's no previous balance to compare against
    (e.g. first row), are left blank rather than flagged -- we only flag
    where we can actually prove a mismatch."""
    if 'Balance' not in df.columns:
        return df, 0

    df = df.copy()
    checks = []
    mismatches = 0
    prev_balance = None

    for _, row in df.iterrows():
        bal = row.get('Balance', np.nan)
        debit = row.get('Debit', np.nan)
        credit = row.get('Credit', np.nan)
        debit = 0.0 if pd.isna(debit) else debit
        credit = 0.0 if pd.isna(credit) else credit

        if prev_balance is None or pd.isna(bal):
            checks.append('')
        else:
            expected = prev_balance + credit - debit
            if abs(expected - bal) > tolerance:
                checks.append('⚠ Mismatch')
                mismatches += 1
            else:
                checks.append('OK')

        if not pd.isna(bal):
            prev_balance = bal

    df['Balance_Check'] = checks
    return df, mismatches


# ----------------------------- PDF <-> Excel conversion -----------------------------

def convert_pdf_to_excel_bytes(uploaded_file, use_ocr=True, force_ocr=False,
                                ocr_resolution=450, apply_bank_correction=False):
    """Raw dump of every extracted PDF row into an in-memory Excel file.
    Returns (BytesIO, row_count)."""
    rows, _ = extract_raw_rows(
        uploaded_file, 'pdf', use_ocr=use_ocr, force_ocr=force_ocr,
        ocr_resolution=ocr_resolution, apply_bank_correction=apply_bank_correction,
    )
    if not rows:
        raise ValueError("No data could be extracted from this PDF.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    max_cols = 0
    for row in rows:
        ws.append(row)
        max_cols = max(max_cols, len(row))

    col_widths = [0] * max_cols
    for row in rows:
        for i, cell in enumerate(row):
            if i < max_cols:
                col_widths[i] = max(col_widths[i], len(str(cell)) if cell is not None else 0)
    for i, width in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = min(max(width + 2, 10), 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, len(rows)


def convert_table_to_pdf_bytes(uploaded_file, ext):
    """Reads an Excel/CSV file and renders it as a formatted in-memory PDF.
    Returns (BytesIO, row_count)."""
    if not REPORTLAB_AVAILABLE:
        raise RuntimeError("reportlab is not installed. Add it to requirements.txt.")

    uploaded_file.seek(0)
    if ext == 'csv':
        df = pd.read_csv(uploaded_file, dtype=str)
    else:
        df = pd.read_excel(uploaded_file, dtype=str)
    df = df.fillna('')

    data = [list(df.columns)] + df.astype(str).values.tolist()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=15 * mm, rightMargin=15 * mm, topMargin=15 * mm, bottomMargin=15 * mm
    )
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(NAVY_BG)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor(GOLD)),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0fdfa')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    doc.build([table])
    buf.seek(0)
    return buf, len(df)


# ----------------------------- Streamlit UI -----------------------------

st.set_page_config(page_title="Bank Statement Keyword Extractor", layout="wide")

st.markdown(
    f"""
    <style>
    .stApp {{ background-color: {NAVY_BG}; }}
    h1, h2, h3 {{ color: {GOLD}; }}
    </style>
    """,
    unsafe_allow_html=True,
)


def login_gate():
    st.title("🔐 Bank Statement Extractor -- Login")
    with st.form("login_form"):
        user = st.text_input("Username")
        pwd = st.text_input("Password", type="password")
        submitted = st.form_submit_button("Login")
    if submitted:
        if user == APP_USERNAME and _hash_password(pwd) == APP_PASSWORD_HASH:
            st.session_state["logged_in"] = True
            st.rerun()
        else:
            st.error("Invalid username or password")
    st.caption(f"Developed by {DEVELOPER_NAME}")


def reset_workflow_state():
    for key in ("raw_rows", "page_images", "ext", "step", "df_raw", "df", "header_idx",
                "row_confidences", "_ocr_pages_used", "_ocr_engine_per_page", "_ocr_was_used"):
        st.session_state.pop(key, None)


def _engine_status_caption():
    engines = []
    engines.append(("PaddleOCR", PADDLEOCR_AVAILABLE))
    engines.append(("Tesseract", PYTESSERACT_AVAILABLE))
    engines.append(("EasyOCR", EASYOCR_AVAILABLE))
    parts = [f"{'✅' if ok else '❌'} {name}" for name, ok in engines]
    st.caption("OCR engines: " + " · ".join(parts) +
               (" · ✅ Camelot table extraction" if CAMELOT_AVAILABLE else " · ❌ Camelot (pdfplumber/OCR only)"))


def keyword_search_tab():
    st.subheader("Upload a bank statement")
    uploaded = st.file_uploader(
        "PDF / Excel / CSV", type=["pdf", "xlsx", "xls", "csv"], key="stmt_upload"
    )

    if not ANY_OCR_ENGINE_AVAILABLE:
        st.caption(
            "ℹ️ No OCR engine is available -- add 'pytesseract' (+ 'tesseract-ocr' in "
            "packages.txt), 'paddleocr'+'paddlepaddle', or 'easyocr' to requirements.txt."
        )
    _engine_status_caption()

    ocr_col1, ocr_col2, ocr_col3 = st.columns(3)
    with ocr_col1:
        use_ocr = st.checkbox(
            "Auto-OCR scanned pages", value=True, disabled=not ANY_OCR_ENGINE_AVAILABLE,
            help="If a PDF page has no readable table (i.e. it's a scanned image), "
                 "run the OCR pipeline on it automatically instead of skipping it.",
        )
    with ocr_col2:
        force_ocr = st.checkbox(
            "Force OCR on every page", value=False,
            disabled=not (ANY_OCR_ENGINE_AVAILABLE and use_ocr),
            help="Use this if a page LOOKS scanned but still returns a messy "
                 "1-column table instead of being auto-detected.",
        )
    with ocr_col3:
        poor_scan = st.checkbox(
            "Poor-quality scan (600 DPI)", value=False,
            disabled=not (ANY_OCR_ENGINE_AVAILABLE and use_ocr),
            help="Higher resolution OCR (600 DPI instead of 450 DPI). Slower, "
                 "but recovers more detail from blurry/low-quality scans.",
        )
    apply_bank_correction = st.checkbox(
        "Auto-correct bank/UPI keywords (GOOGIEPAY -> GOOGLEPAY, etc.)", value=True,
        help="Fixes common OCR misreads of payment-app names using a correction "
             "dictionary, with fuzzy matching for close variants.",
    )
    ocr_resolution = 600 if poor_scan else 450

    if uploaded is not None and st.session_state.get("_last_upload_name") != uploaded.name:
        # New file selected -> reset the wizard state.
        reset_workflow_state()
        st.session_state["_last_upload_name"] = uploaded.name
        ext = uploaded.name.lower().split('.')[-1]
        with st.spinner("Extracting rows... (OCR pages take longer, especially at 600 DPI)"):
            try:
                rows, page_images = extract_raw_rows(
                    uploaded, ext, use_ocr=use_ocr, force_ocr=force_ocr,
                    ocr_resolution=ocr_resolution, apply_bank_correction=apply_bank_correction,
                )
            except Exception as e:
                st.error(f"Failed to read file: {e}")
                return
        if not rows:
            st.error("No data rows could be extracted from this file.")
            return
        st.session_state["raw_rows"] = rows
        st.session_state["page_images"] = page_images
        st.session_state["ext"] = ext
        st.session_state["step"] = 1
        ocr_used = st.session_state.get("_ocr_pages_used") or []
        if ocr_used:
            engine_map = st.session_state.get("_ocr_engine_per_page", {})
            engine_note = ', '.join(f"p{p}:{engine_map.get(p, '?')}" for p in ocr_used)
            st.info(f"🔎 OCR was used on page(s): {engine_note}")

    if "raw_rows" not in st.session_state:
        return

    rows = st.session_state["raw_rows"]
    ext = st.session_state["ext"]

    # ---------------- Step 1: pick header / data-start row ----------------
    if st.session_state.get("step", 1) == 1:
        st.markdown("### Step 1 · Where does your real data start?")
        st.caption(
            "Pick the row number where your Reference / Debit / Credit header actually "
            "begins. Everything above it will be ignored."
        )

        if st.session_state.get("page_images"):
            with st.expander("PDF page preview", expanded=False):
                cols = st.columns(min(4, len(st.session_state["page_images"])))
                for i, img in enumerate(st.session_state["page_images"][:8]):
                    with cols[i % len(cols)]:
                        st.image(img, caption=f"Page {i + 1}", use_container_width=True)

        max_cols = max((len(r) for r in rows), default=0)
        padded_rows = [pad_row(r, max_cols) for r in rows]
        preview_df = pd.DataFrame(
            padded_rows, columns=[f"Col {i + 1}" for i in range(max_cols)]
        )

        # ---- NEW: OCR confidence column + low-confidence flag ----
        confidences = st.session_state.get("row_confidences") or []
        col_config = {}
        if len(confidences) == len(preview_df):
            preview_df["OCR Conf %"] = [round(c, 0) for c in confidences]
            preview_df["⚠"] = ["⚠ low" if c < 60 else "" for c in confidences]
            col_config["OCR Conf %"] = st.column_config.NumberColumn("OCR Conf %", disabled=True)
            col_config["⚠"] = st.column_config.TextColumn("⚠", disabled=True)
            low_conf_count = sum(1 for c in confidences if c < 60)
            if low_conf_count:
                st.warning(
                    f"⚠ {low_conf_count} row(s) have low OCR confidence (<60%). "
                    "Please double-check and correct them below before continuing."
                )

        ocr_used = st.session_state.get("_ocr_pages_used") or []
        if ocr_used:
            st.caption(
                "✏️ **This file used OCR** -- double-click any cell below to fix "
                "misread amounts/text before continuing (e.g. OCR reading '0' as "
                "'O', or a digit wrong)."
            )
        edited_df = st.data_editor(
            preview_df, use_container_width=True, height=320,
            num_rows="fixed", key="raw_rows_editor", column_config=col_config,
        )
        # Push corrections back into `rows` so header selection + everything
        # downstream uses your fixed values, not the raw OCR output. Drop the
        # helper confidence columns before using it as data again.
        edited_data_cols = [c for c in edited_df.columns if c not in ("OCR Conf %", "⚠")]
        data_only_df = edited_df[edited_data_cols]
        rows = data_only_df.astype(object).where(pd.notna(data_only_df), '').values.tolist()
        st.session_state["raw_rows"] = rows

        default_guess = guess_header_row(rows)
        header_idx = st.number_input(
            "Header / data-start row number (from the table above, 0 = first row)",
            min_value=0, max_value=max(len(rows) - 1, 0),
            value=min(default_guess, max(len(rows) - 1, 0)), step=1,
        )

        if st.button("Use this row as header →", type="primary"):
            header_width = len(rows[header_idx])
            header = [str(c).strip() if c is not None else '' for c in rows[header_idx]]

            seen = {}
            clean_header = []
            for i, h in enumerate(header):
                name = h if h else f"Column {i + 1}"
                if name in seen:
                    seen[name] += 1
                    name = f"{name} ({seen[name]})"
                else:
                    seen[name] = 1
                clean_header.append(name)

            auto_map_guess = map_columns(clean_header)
            if 'reference' in auto_map_guess:
                merge_col_idx = clean_header.index(auto_map_guess['reference'])
            else:
                merge_col_idx = header_width - 1

            date_idx = clean_header.index(auto_map_guess['date']) if 'date' in auto_map_guess else None
            debit_idx = clean_header.index(auto_map_guess['debit']) if 'debit' in auto_map_guess else None
            credit_idx = clean_header.index(auto_map_guess['credit']) if 'credit' in auto_map_guess else None

            raw_data_rows = rows[header_idx + 1:]
            row_confs = confidences[header_idx + 1:] if len(confidences) == len(rows) else None
            merged_result = merge_wrapped_continuation_rows(
                raw_data_rows, date_idx=date_idx, debit_idx=debit_idx,
                credit_idx=credit_idx, narration_idx=merge_col_idx, confidences=row_confs
            )
            if row_confs is not None:
                merged_rows, merged_confs = merged_result
                st.session_state["row_confidences_after_header"] = merged_confs
            else:
                merged_rows = merged_result
            data = [pad_row(r, header_width, merge_into=merge_col_idx) for r in merged_rows]

            st.session_state["df_raw"] = pd.DataFrame(data, columns=clean_header)
            st.session_state["step"] = 2
            st.rerun()
        return

    # ---------------- Step 2: confirm columns ----------------
    if st.session_state.get("step") == 2:
        df_raw = st.session_state["df_raw"]
        auto_map = map_columns(df_raw.columns)
        columns = list(df_raw.columns)

        st.markdown("### Step 2 · Confirm your columns")
        st.caption(
            "Auto-detected values are pre-filled below. If your bank uses different "
            "wording (e.g. 'Payment', 'Receipt', 'Recipient'), just change the dropdown."
        )

        c1, c2, c3, c4 = st.columns(4)
        with c1:
            ref_col = st.selectbox(
                "Reference / Narration column", columns,
                index=columns.index(auto_map['reference']) if 'reference' in auto_map else 0,
            )
        with c2:
            debit_options = [NONE_OPTION] + columns
            debit_default = auto_map.get('debit', NONE_OPTION)
            debit_col = st.selectbox(
                "Debit (money out) column", debit_options,
                index=debit_options.index(debit_default) if debit_default in debit_options else 0,
            )
        with c3:
            credit_options = [NONE_OPTION] + columns
            credit_default = auto_map.get('credit', NONE_OPTION)
            credit_col = st.selectbox(
                "Credit (money in) column", credit_options,
                index=credit_options.index(credit_default) if credit_default in credit_options else 0,
            )
        with c4:
            # NEW: optional Balance column -> powers the balance validation engine
            balance_options = [NONE_OPTION] + columns
            balance_default = auto_map.get('balance', NONE_OPTION)
            balance_col = st.selectbox(
                "Balance column (optional, enables balance check)", balance_options,
                index=balance_options.index(balance_default) if balance_default in balance_options else 0,
            )

        st.markdown("**Data preview (first 8 rows)**")
        st.dataframe(df_raw.head(8), use_container_width=True)

        bcol1, bcol2 = st.columns([1, 1])
        with bcol1:
            if st.button("← Back"):
                st.session_state["step"] = 1
                st.rerun()
        with bcol2:
            if st.button("Build table →", type="primary"):
                if debit_col == NONE_OPTION and credit_col == NONE_OPTION:
                    st.warning("Select at least one of Debit or Credit column.")
                elif debit_col == credit_col and debit_col != NONE_OPTION:
                    st.warning("Debit and Credit cannot be the same column.")
                else:
                    col_map = {'reference': ref_col}
                    if debit_col != NONE_OPTION:
                        col_map['debit'] = debit_col
                    if credit_col != NONE_OPTION:
                        col_map['credit'] = credit_col
                    if balance_col != NONE_OPTION:
                        col_map['balance'] = balance_col

                    ocr_mode = bool(st.session_state.get("_ocr_was_used"))
                    df = build_clean_df(df_raw, col_map, ocr_mode=ocr_mode)

                    # NEW: run balance validation if a Balance column was mapped
                    mismatches = 0
                    if 'balance' in col_map:
                        df, mismatches = validate_balances(df)

                    st.session_state["df"] = df
                    st.session_state["balance_mismatches"] = mismatches
                    st.session_state["step"] = 3
                    st.rerun()
        return

    # ---------------- Step 3: keyword search ----------------
    if st.session_state.get("step") == 3 and "df" in st.session_state:
        df = st.session_state["df"]
        st.success(f"Table ready · {len(df)} rows")

        # NEW: balance validation warning banner
        mismatches = st.session_state.get("balance_mismatches", 0)
        if 'Balance_Check' in df.columns:
            if mismatches > 0:
                st.warning(
                    f"⚠ Balance check found {mismatches} row(s) where "
                    "Previous Balance + Credit - Debit doesn't match the stated Balance. "
                    "These are marked '⚠ Mismatch' in the Balance_Check column below -- "
                    "likely an OCR misread, please verify."
                )
                with st.expander("Show only mismatched rows"):
                    st.dataframe(df[df['Balance_Check'] == '⚠ Mismatch'], use_container_width=True)
            else:
                st.success("✅ Balance check passed on all rows -- Previous Balance + Credit - Debit matches throughout.")

        with st.expander("✏️ Fix wrong amounts before searching (click any cell)", expanded=False):
            st.caption(
                "OCR can misread digits (e.g. 5000 read as 5OOO, or a wrong "
                "decimal). Edit any Reference/Debit/Credit/Balance value below -- "
                "your changes are used for the search and totals."
            )
            col_config = {}
            if 'Debit' in df.columns:
                col_config['Debit'] = st.column_config.NumberColumn("Debit", format="%.2f")
            if 'Credit' in df.columns:
                col_config['Credit'] = st.column_config.NumberColumn("Credit", format="%.2f")
            if 'Balance' in df.columns:
                col_config['Balance'] = st.column_config.NumberColumn("Balance", format="%.2f")
            if 'Balance_Check' in df.columns:
                col_config['Balance_Check'] = st.column_config.TextColumn("Balance_Check", disabled=True)
            df = st.data_editor(
                df, use_container_width=True, height=320,
                column_config=col_config, key="final_amounts_editor",
            )
            st.session_state["df"] = df

        available_cols = [c for c in ['Debit', 'Credit'] if c in df.columns]

        s1, s2, s3 = st.columns([2, 1, 1])
        with s1:
            keyword = st.text_input("Keyword")
        with s2:
            col = st.selectbox("Column", available_cols)
        with s3:
            st.write("")
            st.write("")
            search_clicked = st.button("🔍 Search", type="primary")

        if st.button("↺ Start over with a new file"):
            reset_workflow_state()
            st.rerun()

        if search_clicked:
            if not keyword.strip():
                st.warning("Enter a keyword to search for.")
            else:
                mask = df['Reference'].str.contains(keyword, case=False, na=False, regex=False)
                matches = df[mask]
                amounts = matches[col].dropna()
                amounts = amounts[amounts > 0]

                if amounts.empty:
                    st.info(f"No '{col}' entries found for keyword '{keyword}'.")
                else:
                    amt_str = ''.join(
                        [f"+{int(a) if float(a) == int(a) else a}" for a in amounts]
                    )
                    st.text_area("Amounts", amt_str, height=100)
                    st.markdown(f"**Found {len(amounts)} entries · Total: ₹{amounts.sum():,.2f}**")

                    result_table = matches[matches[col] > 0][['Reference', col]]
                    st.dataframe(result_table, use_container_width=True)

                    csv_bytes = result_table.to_csv(index=False).encode('utf-8')
                    st.download_button(
                        "Download matches as CSV", data=csv_bytes,
                        file_name="matches.csv", mime="text/csv",
                    )


def converter_tab():
    st.subheader("PDF → Excel")
    pdf_file = st.file_uploader("Select PDF to convert", type=["pdf"], key="pdf2xlsx")
    c_ocr1, c_ocr2, c_ocr3 = st.columns(3)
    with c_ocr1:
        conv_use_ocr = st.checkbox(
            "Auto-OCR scanned pages", value=True, disabled=not ANY_OCR_ENGINE_AVAILABLE,
            key="conv_use_ocr",
        )
    with c_ocr2:
        conv_force_ocr = st.checkbox(
            "Force OCR on every page", value=False,
            disabled=not (ANY_OCR_ENGINE_AVAILABLE and conv_use_ocr), key="conv_force_ocr",
        )
    with c_ocr3:
        conv_poor_scan = st.checkbox(
            "Poor-quality scan (600 DPI)", value=False,
            disabled=not (ANY_OCR_ENGINE_AVAILABLE and conv_use_ocr), key="conv_poor_scan",
        )
    conv_bank_correction = st.checkbox(
        "Auto-correct bank/UPI keywords", value=True, key="conv_bank_correction",
    )
    if pdf_file is not None and st.button("Convert to Excel"):
        try:
            with st.spinner("Converting... (OCR pages take longer)"):
                buf, row_count = convert_pdf_to_excel_bytes(
                    pdf_file, use_ocr=conv_use_ocr, force_ocr=conv_force_ocr,
                    ocr_resolution=600 if conv_poor_scan else 450,
                    apply_bank_correction=conv_bank_correction,
                )
            st.success(f"Converted successfully! {row_count} rows.")
            st.download_button(
                "Download Excel file", data=buf,
                file_name=pdf_file.name.rsplit('.', 1)[0] + ".xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as e:
            st.error(f"Conversion failed: {e}")

    st.divider()

    st.subheader("Excel/CSV → PDF")
    if not REPORTLAB_AVAILABLE:
        st.warning("reportlab is not installed -- add it to requirements.txt to enable this tool.")
    table_file = st.file_uploader(
        "Select Excel/CSV to convert", type=["xlsx", "xls", "csv"], key="tbl2pdf"
    )
    if table_file is not None and st.button("Convert to PDF", disabled=not REPORTLAB_AVAILABLE):
        ext = table_file.name.lower().split('.')[-1]
        try:
            with st.spinner("Converting..."):
                buf, row_count = convert_table_to_pdf_bytes(table_file, ext)
            st.success(f"Converted successfully! {row_count} rows.")
            st.download_button(
                "Download PDF file", data=buf,
                file_name=table_file.name.rsplit('.', 1)[0] + ".pdf",
                mime="application/pdf",
            )
        except Exception as e:
            st.error(f"Conversion failed: {e}")


def main_app():
    st.title("Bank Statement Keyword Amount Extractor")
    st.caption("Search transaction amounts by keyword · Convert PDF ⇄ Excel")

    tab1, tab2 = st.tabs(["🔍 Keyword Search", "🔁 Format Converter"])
    with tab1:
        keyword_search_tab()
    with tab2:
        converter_tab()

    st.divider()
    st.caption(f"Developed by {DEVELOPER_NAME}")


if "logged_in" not in st.session_state:
    st.session_state["logged_in"] = False

if not st.session_state["logged_in"]:
    login_gate()
else:
    main_app()
